from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from base.models import Material
from base.models import Purchase
from base.models import Need
from django.contrib import messages
from django.core.files.storage import default_storage
from django.contrib.auth.hashers import check_password
from django.core.exceptions import ValidationError
from django.utils.timezone import now
from django.utils.dateparse import parse_date
from django.core.mail import send_mail
import openpyxl
import os
import csv
import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import io
import urllib, base64
from io import BytesIO
from django.http import HttpResponse,Http404
from django.db import IntegrityError
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,Paragraph, Spacer,Image
from django.utils import timezone
from django.db import connection
from django.core.mail import send_mail
from reportlab.lib.units import inch
from .models import StaffUsers
from django.contrib.auth.hashers import check_password 
from .models import Mess_menu_Schedule
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from .models import Purchase, Received
from django.http import JsonResponse
from django.db.models import Sum
from .models import Item
from openpyxl import load_workbook
from django.db.models import OuterRef, Subquery
from django.core.mail import send_mail
from django.urls import reverse
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.template.loader import render_to_string
from django.utils.crypto import get_random_string
import random
from django.core.mail import EmailMultiAlternatives
from decimal import Decimal
from django.db.models import Sum, Subquery, OuterRef
from matplotlib.colors import to_hex
import seaborn as sns





def home(request):
    return render(request, 'base/home.html')

# view for the login page
def login_view(request):
    if request.method == 'POST':
        staff_id = request.POST['staff_id'].strip()  # Strip any leading/trailing whitespace
        password = request.POST['password'].strip()  # Strip any leading/trailing whitespace

        # Attempt to get the user from the StaffUsers model
        user = StaffUsers.objects.filter(staff_id=staff_id).first()  # Get the first matching user

        if user:  # Check if user exists
            # Use check_password to compare the plaintext password to the hashed password
            if check_password(password, user.password):  # Compare plaintext password to the hashed password
                if not user.is_admin:  # Ensure no admin login as staff
                    login(request, user)  # Assuming you have a login method for your custom user
                    
                    # Store the user's role in the session
                    request.session['role'] = user.role  # Save the role in the session

                    return redirect('dash_view')  # Redirect to a common dashboard page
                else:
                    messages.error(request, 'Admin login is not allowed.')
            else:
                messages.error(request, 'Invalid password.')
        else:
            messages.error(request, 'Invalid staff ID.')

    return render(request, 'base/home.html')


def generate_otp():
    return str(random.randint(100000, 999999))

# View for the OTP and Mail message Page
def password_reset_request(request):
    if request.method == 'POST':
        email = request.POST['email']
        user = StaffUsers.objects.filter(email=email).first()
        if user:
            otp = generate_otp()
            user.otp = otp
            user.otp_expiration = timezone.now() + timedelta(minutes=10)  # OTP valid for 10 minutes
            user.save()
            
            # Store the email in session
            request.session['reset_email'] = email
            
            # Create a descriptive email message with OTP
            subject = 'Password Reset Request'
            text_content = f"Your OTP for password reset is: {otp}. This OTP is valid for 10 minutes."
            html_content = f"""
                <p>Hello,</p>
                <p>We received a request to reset your password. To proceed, please use the OTP provided below. 
                This OTP is valid for 10 minutes.</p>
                <p><strong>OTP:</strong> <strong style="font-size:18px;">{otp}</strong></p>
                <p>If you did not request a password reset, you can ignore this email.</p>
                <p>Thank you!</p>
            """
            
            # Send the email with HTML formatting
            email_message = EmailMultiAlternatives(subject, text_content, settings.DEFAULT_FROM_EMAIL, [email])
            email_message.attach_alternative(html_content, "text/html")
            email_message.send(fail_silently=False)

            messages.success(request, 'OTP has been sent to your email.')
            return redirect('verify_otp')
        else:
            messages.error(request, 'Email not found.')
    
    return render(request, 'base/email.html')

# View for the OTP verification Page

def verify_otp(request):
    if request.method == 'POST':
        otp = request.POST['otp']
        email = request.session.get('reset_email')  # Retrieve email from session
        user = StaffUsers.objects.filter(email=email, otp=otp).first()
        
        if user and user.otp_expiration >= timezone.now():
            user.otp = None  # Clear OTP after successful verification
            user.otp_expiration = None
            user.save()
            
            # Store staff_id and role in session for password reset page
            request.session['staff_id'] = user.staff_id
            request.session['role'] = user.role
            
            return redirect('password_reset_form')
        else:
            messages.error(request, 'Invalid or expired OTP.')
    
    return render(request, 'base/otp_verify.html')

# View to handle password reset form
def password_reset_form(request):
    if request.method == 'POST':
        new_password = request.POST['new_password']
        confirm_password = request.POST['confirm_password']
        
        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect('password_reset_form')

        # Retrieve staff_id from session
        staff_id = request.session.get('staff_id')
        user = StaffUsers.objects.filter(staff_id=staff_id).first()
        
        if user:
            user.set_password(new_password)
            user.save()
            messages.success(request, "Your password has been successfully reset.")
            
            # Clear session data after password reset
            request.session.pop('staff_id', None)
            request.session.pop('role', None)
            request.session.pop('reset_email', None)  # Clear reset email
            
            return redirect('home')  # Redirect to login page
        else:
            messages.error(request, "User not found.")
    
    return render(request, 'base/password.html')


# View for the Sidebar Page
@login_required
def sidebar_view(request):
    if request.method == 'POST':
        staff_id = request.POST['staff_id'].strip()
        user_role = request.session.get('role', 'Guest')
        return render(request,'base/sidebar.html')

# View for the Available Stock Page
@login_required
def dash_view(request):
    user_role = request.session.get('role', 'Guest')  # Default to 'Guest' if role is not set
    selected_names = request.GET.getlist('names')  # This will return a list of selected names
    selected_month = request.GET.get('month', '')
    selected_date = request.GET.get('selected_date', '')

    # Get all data entries
    data_entries = Material.objects.all()

    # Filter by selected names if provided
    if selected_names:
        data_entries = data_entries.filter(name__in=selected_names)

    # Filter by selected month if provided
    if selected_month:
        data_entries = data_entries.filter(date__month=selected_month)

    # Filter by selected date if provided
    if selected_date:
        data_entries = data_entries.filter(date=selected_date)

    # Get distinct material names for the search box
    all_materials = Material.objects.values_list('name', flat=True).distinct()

    # Pass filtered data, materials list, and user role to the template
    context = {
        'data_entries': data_entries,
        'all_materials': all_materials,
        'selected_names': selected_names,
        'selected_month': selected_month,
        'selected_date': selected_date,
        'user_role': user_role,  # Add the user's role to the context
    }
    
    return render(request, 'base/port.html', context)

#Available stock pdf
@login_required
def download_pdf(request):
    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="data.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=A4)

    # Container for the 'Flowable' objects (tables, paragraphs, etc.)
    elements = []

    # Add the image to the PDF
    logo_path = r"C:\Users\sdhan\OneDrive\Desktop\Projects\Hostel mess\base\static\images\RIT_logo.png"  # Use the provided image path
    logo = Image(logo_path, width=1.2*inch, height=1.2*inch)  # Adjust width and height as needed

    # Add the title to the PDF
    styles = getSampleStyleSheet()
    title = Paragraph("<b>RAMCO INSTITUTE OF TECHNOLOGY RAJAPALAYAM</b>", styles['Title'])

    # Create a table for the image and title side by side in one row
    # The table will have two columns: one for the image, and one for the title
    header_table = Table([[logo, title]], colWidths=[1.5*inch, 5.5*inch])  # Adjust column widths as needed

    # Set the style for the table
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),  # Align the logo to the left
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),  # Align the title to the left
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Align both the image and title vertically in the middle
        ('LEFTPADDING', (1, 0), (1, 0), 15),  # Add some space between the image and the title
        ('TOPPADDING', (0, 0), (-1, -1), 10),  # Add top padding
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),  # Add bottom padding
    ]))

    # Add the header (image + title) to the elements list
    elements.append(header_table)

    # Define table data, starting with headers
    data = [['Date', 'Name', 'Quantity', 'Available Quantity', 'Used Quantity', 'Rate', 'Amount']]

    # Fetch data from your model
    entries = Material.objects.all()

    for entry in entries:
        data.append([
            str(entry.date or ''),
            str(entry.name or ''),
            str(entry.quantity if entry.quantity is not None else ''),
            str(entry.available_quantity if entry.available_quantity is not None else ''),
            str(entry.used_quantity if entry.used_quantity is not None else ''),
            str(entry.rate if entry.rate is not None else ''),
            str(entry.amount if entry.amount is not None else '')
        ])

    # Create a table with the data
    table = Table(data)

    # Add a table style to adjust alignment and add lines (no colors)
    style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align all text in the table to the center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for headers
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines (no background colors)
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Body font
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Body font size
    ])

    table.setStyle(style)

    # Add the table to the elements list
    elements.append(table)

    # Build the PDF
    doc.build(elements)

    return response


#adding purchased data using data.html
@login_required
def add_data(request):
    user_role = request.session.get('role', 'Guest')  # Fetch the user role from session
    
    if request.method == 'POST':
        # Extracting data from POST request
        material_dates = request.POST.getlist('material-date')
        material_names = request.POST.getlist('material-name')
        material_quantities = request.POST.getlist('material-quantity')
        material_units = request.POST.getlist('material-unit')
        material_rates = request.POST.getlist('material-rate')
    
        try:
            for i in range(len(material_names)):
                # Parse the input data
                date = parse_date(material_dates[i])
                name = material_names[i]
                quantity = float(material_quantities[i]) if material_quantities[i] else 0
                quantity=Decimal(quantity)
                unit = material_units[i]
                rate = Decimal(material_rates[i]) if material_rates[i] else Decimal(0)

                # Unit conversion logic
                if unit == 'g':  # Grams to KGS
                    quantity /= 1000
                    unit = 'KGS'
                elif unit == 'mL':  # Milliliters to LTR
                    quantity /= 1000
                    unit = 'LTR'
                elif unit == 'kg':  # Kilograms to KGS
                    unit = 'KGS'
                elif unit == 'l':  # Liters to LTR
                    unit = 'LTR'

                amount = quantity * Decimal(rate)

                # Fetch any existing entry for this material on the same date
                existing_entry = Material.objects.filter(name=name, date=date,unit=unit).first()

                if existing_entry:
                    # Update the existing row if it's the same day and material
                    existing_entry.quantity += quantity
                    existing_entry.available_quantity += quantity
                    existing_entry.unit = unit or existing_entry.unit  # Keep previous value if unit is empty
                    existing_entry.rate = rate or existing_entry.rate  # Keep previous value if rate is empty
                    existing_entry.amount = amount or existing_entry.amount  # Update amount
                    existing_entry.save()
                else:
                    # Fetch the last entry for this material before the current date
                    last_entry = Material.objects.filter(name=name, date__lt=date).order_by('-date').first()

                    if last_entry:
                        available_quantity = last_entry.available_quantity + quantity
                    else:
                        available_quantity = quantity

                    # Create a new entry for the material
                    Material.objects.create(
                        date=date,
                        name=name,
                        quantity=quantity,
                        used_quantity=0,  # No used quantity for purchase
                        available_quantity=available_quantity,
                        unit=unit,
                        rate=rate,
                        amount=amount
                    )

                     # Mark the received data as fetched
                Received.objects.filter(received_date=date, name=name).update(is_fetched=True)

            messages.success(request, "Purchase data successfully stored!")

        except Exception as e:
            # Log the error for debugging if needed
            print(f"Error storing data: {e}")
            messages.error(request, f"Data not stored! Error: {str(e)}")

        return redirect('dash_view')

    context = {
        'user_role': user_role,  # Add the user's role to the context
    }

    return render(request, 'base/data.html', context)


#adding usage data using used.html

@login_required
def adddata(request):
    user_role = request.session.get('role', 'Guest')
    
    if request.method == 'POST':
        # Fetching lists of data from the POST request
        material_dates = request.POST.getlist('material-date')
        material_names = request.POST.getlist('material-name')
        material_quantities = request.POST.getlist('material-quantity')
        material_units = request.POST.getlist('material-unit')

        try:
            # Iterating through each material entry submitted
            for i in range(len(material_names)):
                date = parse_date(material_dates[i])  # Parse the date
                name = material_names[i]  # Material name
                unit = material_units[i].lower()  # Unit (for display or calculation)
                used_quantity = float(material_quantities[i]) if material_quantities[i] else 0  # Used quantity
                used_quantity = Decimal(used_quantity)  # Convert to Decimal

                # Convert quantities to base units (KGS for weight, MLTR for volume)
                if unit == 'g':  # Grams to KGS
                    used_quantity /= 1000
                    unit = 'KGS'  # Standardize unit to KGS
                elif unit == 'ml':  # Milliliters to LTR
                    used_quantity /= 1000
                    unit = 'MLTR'  # Standardize unit to MLTR
                elif unit == 'kg':  # Kilograms to KGS
                    unit = 'KGS'  # Standardize unit to KGS
                elif unit == 'l':  # Liters remain LTR without conversion
                    unit = 'LTR'  # Standardize unit to LTR
                elif unit == 'nos':  # No conversion needed for NOS (Number of pieces)
                    unit = 'NOS'  # Ensure unit is standardized as NOS
                elif unit == 'pac':  # No conversion needed for PAC (Packs)
                    unit = 'PAC'  # Ensure unit is standardized as PAC
                elif unit == 'btl':  # No conversion needed for BTL (Bottles)
                    unit = 'BTL'  # Ensure unit is standardized as BTL
                elif unit == 'mtr':  # No conversion needed for MTR (Meters)
                    unit = 'MTR'  # Ensure unit is standardized as MTR

                # Fetching the existing material entry for the same material and date
                existing_entry = Material.objects.filter(name=name, date=date, unit=unit).first()

                if existing_entry:
                    # Ensure the existing entry is in the same standardized unit
                    # Only check for unit mismatch if it's one of the units we are converting (KGS or MLTR)
                    if unit in ['KGS', 'MLTR'] and existing_entry.unit != unit:
                        raise ValueError(f"Unit mismatch for material '{name}'. Ensure consistent units (e.g., KGS/GS or LTR/MLTR).")

                    # Update the existing entry (same date, same material)
                    existing_entry.used_quantity += used_quantity  # Adding as Decimal
                    existing_entry.available_quantity = max(existing_entry.available_quantity - used_quantity, Decimal(0))  # Use Decimal for comparison
                    existing_entry.save()
                else:
                    # Fetching the last entry for the material to carry forward stock info
                    last_entry = Material.objects.filter(name=name).order_by('-date').first()

                    # Logic for available quantity
                    if last_entry:
                        # Only raise error if there is a unit mismatch with a unit we are converting (KGS or MLTR)
                        if last_entry.unit in ['KGS', 'MLTR'] and last_entry.unit != unit:
                            raise ValueError(f"Unit mismatch for material '{name}'. Ensure consistent units (e.g., KGS/GS or LTR/MLTR).")
                        available_quantity = max(last_entry.available_quantity - used_quantity, Decimal(0))  # Use Decimal for comparison
                    else:
                        available_quantity = -used_quantity  # If no previous entry, set negative quantity

                    # Create a new material entry
                    material = Material.objects.create(
                        date=date,
                        name=name,
                        quantity=0,  # No purchase quantity for usage
                        used_quantity=used_quantity,
                        available_quantity=available_quantity,
                        unit=unit,  # Store the standardized unit
                        rate=last_entry.rate if last_entry else None,  # Carry forward rate from last entry if available
                        amount=last_entry.amount if last_entry else None  # Carry forward amount if available
                    )
                    material.save()
                     # Mark corresponding Need entry as fetched when usage is entered
                Need.objects.filter(date=date, name=name).update(is_fetched=True)

            # Success message
            messages.success(request, "Usage data successfully stored!")

        except Exception as e:
            # If any exception occurs, show an error message
            messages.error(request, f"Data not stored! Error: {str(e)}")
    
        return redirect('dash_view')  # Redirect after processing

    context = {
        'user_role': user_role,  # Pass user role for visibility or access control
    }

    return render(request, 'base/used.html', context)  # Render the used.html template



#approval page using acknowledge.html
@login_required
def Acknowledgement_view(request):
    user_role = request.session.get('role', 'Guest')
    pending_purchase_count = Purchase.objects.filter(is_acknowledged=False).count()
    
    # Count unapproved needs
    pending_need_count = Need.objects.filter(is_acknowledged=False).count()

    context = {
        'pending_purchase_count': pending_purchase_count,
        'pending_need_count': pending_need_count,
        'user_role': user_role, 

    }
    return render(request, 'base/acknowledge.html', context)

#Approval for shopkeeper using purchase_acknowledge.html
@login_required
def purchase_Acknowledgement_view(request):
    user_role = request.session.get('role', 'Guest')

    if request.method == "POST":
        # Get unacknowledged materials
        materials = Purchase.objects.filter(is_acknowledged=False)

        if materials.exists():
            materials.update(is_acknowledged=True)  # Mark them as acknowledged
            
            # Prepare the email to notify the other user
            email_subject = "Materials Acknowledged"
            email_message = (
                            "Materials have been approved to purchase .\n\n"
                             "Please visit the page at your earliest convenience to review and print the material to purchase .\n\n"
                              "Thank you for your attention to this matter")
            
            recipient_list = ['sdhanesh09829@gmail.com']  # Change to the recipient's email
            
            # Sending email with error handling
            try:
                send_mail(
                    subject=email_subject,
                    message=email_message,
                    from_email='ehostelrit@gmail.com',
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
                messages.success(request, "Materials acknowledged successfully and notification sent!")
            except Exception as e:
                messages.error(request, f'Error sending email: {str(e)}')
        else:
            messages.info(request, "No materials to acknowledge.")

        return redirect('Purchase_Acknowledgement')  # Redirect to the same view to show acknowledged materials

    # Fetch unacknowledged materials for display
    unacknowledged_materials = Purchase.objects.filter(is_acknowledged=False)

    return render(request, 'base/purchase_acknowledge.html', {
        'data_entries': unacknowledged_materials,
        'user_role': user_role,
    })

#Approval for Mess-incharge using need_acknowledge.html
@login_required
def need_Acknowledgement_view(request):
    user_role = request.session.get('role', 'Guest')

    if request.method == "POST":
        # Get unacknowledged materials
        materials = Need.objects.filter(is_acknowledged=False)

        if materials.exists():
            materials.update(is_acknowledged=True)  # Mark them as acknowledged
            
            # Prepare the email to notify the other user
            email_subject = "Materials Acknowledged"
            email_message = ("Materials have been approved to use the stock items  .\n\n"
                             "Please visit the page at your earliest convenience to take print and use the approved materials.\n\n"
                             "Thank you for your attention to this matter")

            recipient_list = ['sdhanesh09829@gmail.com']  # Change to the recipient's email
            
            # Sending email with error handling
            try:
                send_mail(
                    subject=email_subject,
                    message=email_message,
                    from_email='ehostelrit@gmail.com',
                    recipient_list=recipient_list,
                    fail_silently=False,
                )
                messages.success(request, "Materials acknowledged successfully and notification sent!")
            except Exception as e:
                messages.error(request, f'Error sending email: {str(e)}')
        else:
            messages.info(request, "No materials to acknowledge.")

        return redirect('need_Acknowledgement')  # Redirect to the same view to show acknowledged materials

    # Fetch unacknowledged materials for display
    unacknowledged_materials = Need.objects.filter(is_acknowledged=False)

    return render(request, 'base/need_acknowledge.html', {
        'data_entries': unacknowledged_materials,
        'user_role': user_role,
    })

#mess schedule page view for mess.html
@login_required
def mess_menu(request):
    user_role = request.session.get('role', 'Guest')
    day = request.GET.get('form-day')
    year = request.GET.get('form-year')
    menus = []

    print("Day:", day)
    print("Year:", year)
    print("menu:", menus)

    # Fetch the meal schedule based on the selected day and year
    if day and year:
        try:
            menus = Mess_menu_Schedule.objects.filter(day=day, year=year)  # Adjust this query based on your model fields
        except Exception as e:
            print(f"Error fetching menus: {e}")

    context = {
        'day': day,
        'year': year,
        'menus': menus,
        'user_role': user_role,
    }
    
    return render(request, 'base/mess.html', context)

#print.html 
@login_required
def print_view(request):
    user_role = request.session.get('role', 'Guest')



    context = {
        'user_role': user_role, 
       # Add the user's role to the context
    }
    return render(request,'base/print.html',context)


#shopkeeper Approval print page using purchase_print.html
@login_required
def print_purchase_view(request):
       user_role = request.session.get('role', 'Guest')
       data_entries = Purchase.objects.filter(is_acknowledged=True,is_printed=False)
       if request.method == "POST": 
           data_entries.update(is_printed=True)
       return render(request, 'base/purchase_print.html', {'data_entries': data_entries, 'user_role': user_role})

#mess-incharge Approval print page using need_print.html
@login_required
def print_need_view(request):
       user_role = request.session.get('role', 'Guest')
       data_entries = Need.objects.filter(is_acknowledged=True,is_printed=False)
       if request.method == "POST": 
           data_entries.update(is_printed=True)
       return render(request, 'base/need_print.html', {'data_entries': data_entries ,'user_role': user_role})

#shopkeeper print(purchase)
def print_pdf(request):
    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    unique_reference_count = Purchase.objects.exclude(reference_id__isnull=True).values('reference_id').distinct().count()
    
    # Check if there are any entries
    entries = Purchase.objects.filter(is_acknowledged=True, is_printed=False)

    if not entries.exists():
         messages.error(request, 'There is No Data to Print !')  # Check if there are no entries
        # Render an error message on a new HTML page
         return redirect('print_purchase_view')  # Return error template instead of PDF

    response['Content-Disposition'] = f'attachment; filename="rit/mess/purchase/{unique_reference_count+1}.pdf"'

    # Create the PDF object 
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    title_style = styles['Heading1']

    # Add the logo and institution name in a table to align them horizontally
    logo_path = r'C:\Users\sdhan\OneDrive\Desktop\new\base\static\images\RIT_logo.png'  # Your logo's path
    logo = Image(logo_path)
    logo.drawHeight = 1.25 * inch  # Adjust height
    logo.drawWidth = 1.25 * inch  # Adjust width

    # Create a table with the logo and institution name side by side
    title_table_data = [
        [logo, Paragraph('RAMCO INSTITUTE OF TECHNOLOGY RAJAPALAYAM', title_style)]
    ]
    title_table = Table(title_table_data, colWidths=[1.5 * inch, 4.5 * inch])  # Adjust column widths as needed
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),  # Vertically align logo in the middle
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),  # Vertically align text in the middle
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 20))

    # Add placeholders for supplier and date
    elements.append(Paragraph('To : ___________                 Date : _______________ ', normal_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph('       M/s  :  _________________________', normal_style))
    elements.append(Spacer(1, 15))

    # Instructional text
    elements.append(Paragraph('Please supply the following list of items:', normal_style))
    elements.append(Spacer(1, 12))

    # Define table data, starting with headers
    data = [['S.No.', 'Date', 'Name', 'Required Quantity','unit']]

    count = 1
    reference_id = f"rit/mess/purchase/{unique_reference_count+1}"

    for entry in entries:
        data.append([
            count,
            str(entry.date or ''),
            str(entry.name or ''),
            str(entry.quantity if entry.quantity is not None else ''),
            str(entry.unit or '')
        ])
        # Update reference_id and mark as printed
        entry.reference_id = reference_id
        entry.is_printed = True
        entry.save()
        count += 1

    # Create the table without colors
    table = Table(data, colWidths=[80, 120, 150,120,80])  # Adjust column widths

    # Add table style (No color styling)
    table_style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align text to center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for header
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Body font
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Body font size
    ])
    table.setStyle(table_style)

    # Add the table to elements
    elements.append(table)
    elements.append(Spacer(1, 24))

    # Footer: Requested by, Accountant, GM(A) with enough space for signatures
    footer_table_data = [
        ['Requested by: ', '', 'Accountant: ', '', 'GM(A): ']
    ]
    footer_table = Table(footer_table_data, colWidths=[80, 80, 80, 80, 80])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT')  # Align text to the left
    ]))
    elements.append(footer_table)

    # Build the PDF
    doc.build(elements)

    return response


#mess-incharge print (Usage)
def need_print_pdf(request):
    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    unique_reference_count = Need.objects.exclude(reference_id__isnull=True).values('reference_id').distinct().count()
    
    # Fetch data from your model
    entries = Need.objects.filter(is_acknowledged=True, is_printed=False)

    if not entries.exists():  
         messages.error(request,'There is No Data to Print!')# Check if there are no entries
        # Render an error message on a new HTML page
         return redirect('print_need_view')  # Return error template instead of PDF

    response['Content-Disposition'] = f'attachment; filename="rit/mess/need/{unique_reference_count + 1}.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    title_style = styles['Heading1']

    # Add the logo and institution name in a table to align them horizontally
    logo_path = r'C:\Users\sdhan\OneDrive\Desktop\Projects\Hostel mess\base\static\images\RIT_logo.png'  # Your logo's path
    logo = Image(logo_path)
    logo.drawHeight = 1.25 * inch  # Adjust height
    logo.drawWidth = 1.25 * inch  # Adjust width

    # Create a table with the logo and institution name side by side
    title_table_data = [
        [logo, Paragraph('RAMCO INSTITUTE OF TECHNOLOGY RAJAPALAYAM', title_style)]
    ]
    title_table = Table(title_table_data, colWidths=[1.5 * inch, 4.5 * inch])  # Adjust column widths as needed
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),  # Vertically align logo in the middle
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),  # Vertically align text in the middle
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 20))

    # Add placeholders for supplier and date
    elements.append(Paragraph('To : ___________                 Date : _______________ ', normal_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph('       M/s  :  _________________________', normal_style))
    elements.append(Spacer(1, 15))

    # Instructional text
    elements.append(Paragraph('Please supply the following list of items:', normal_style))
    elements.append(Spacer(1, 12))

    # Define table data (headers)
    data = [['S.No.', 'Date', 'Name of the item', 'Required Quantity','unit']]

    count = 1
    reference_id = f"rit/mess/Need/{unique_reference_count + 1}"

    # Populate the table with data
    for entry in entries:
        data.append([
            count,
            str(entry.date or ''),  # Incremental number
            str(entry.name or ''),  # Name of the item
            str(entry.quantity if entry.quantity is not None else '') , # Quantity
            str(entry.unit or'')
        ])

        # Assign reference_id and mark as printed
        entry.reference_id = reference_id
        entry.is_printed = True
        entry.save()

        count += 1

    # Create the table without colors
    table = Table(data, colWidths=[80, 120, 150,120,80])  # Adjust column widths to match

    # Add table style (No color styling)
    table_style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align text to center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for header
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Body font
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Body font size
    ])
    table.setStyle(table_style)

    # Add the table to elements
    elements.append(table)
    elements.append(Spacer(1, 24))

    # Footer: Requested by, Accountant, GM(A) with enough space for signatures
    footer_table_data = [
        ['Requested by: ', '', 'Accountant: ', '', 'GM(A): ']
    ]
    footer_table = Table(footer_table_data, colWidths=[80, 80, 80, 80, 80])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT')  # Align text to the left
    ]))
    elements.append(footer_table)

    # Build the PDF
    doc.build(elements)

    return response

#backup for Purchse print

def backup_purchase_print_pdf(request):
    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    unique_reference_count = Purchase.objects.exclude(reference_id__isnull=True).values('reference_id').distinct().count()
    
    # Check if there are any entries
    entries = Purchase.objects.filter(is_acknowledged=True, is_printed=True)

    if not entries.exists():
         messages.error(request, 'There is No Data to Print !')  # Check if there are no entries
        # Render an error message on a new HTML page
         return redirect('backup_purchase_view')  # Return error template instead of PDF

    response['Content-Disposition'] = f'attachment; filename="rit/mess/purchase/{unique_reference_count+1}.pdf"'

    # Create the PDF object 
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    title_style = styles['Heading1']

    # Add the logo and institution name in a table to align them horizontally
    logo_path = r'C:\Users\sdhan\OneDrive\Desktop\Projects\Hostel mess\base\static\images\RIT_logo.png'  # Your logo's path
    logo = Image(logo_path)
    logo.drawHeight = 1.25 * inch  # Adjust height
    logo.drawWidth = 1.25 * inch  # Adjust width

    # Create a table with the logo and institution name side by side
    title_table_data = [
        [logo, Paragraph('RAMCO INSTITUTE OF TECHNOLOGY RAJAPALAYAM', title_style)]
    ]
    title_table = Table(title_table_data, colWidths=[1.5 * inch, 4.5 * inch])  # Adjust column widths as needed
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),  # Vertically align logo in the middle
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),  # Vertically align text in the middle
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 20))

    # Add placeholders for supplier and date
    elements.append(Paragraph('To : ___________                 Date : _______________ ', normal_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph('       M/s  :  _________________________', normal_style))
    elements.append(Spacer(1, 15))

    # Instructional text
    elements.append(Paragraph('Please supply the following list of items:', normal_style))
    elements.append(Spacer(1, 12))

    # Define table data, starting with headers
    data = [['S.No.', 'Date', 'Name', 'Required Quantity','unit']]

    count = 1
    reference_id = f"rit/mess/purchase/{unique_reference_count+1}"

    for entry in entries:
        data.append([
            count,
            str(entry.date or ''),
            str(entry.name or ''),
            str(entry.quantity if entry.quantity is not None else ''),
            str(entry.unit or '')
        ])
        # Update reference_id and mark as printed
        entry.reference_id = reference_id
        entry.is_printed = True
        entry.save()
        count += 1

    # Create the table without colors
    table = Table(data, colWidths=[80, 120, 150,120,80])  # Adjust column widths

    # Add table style (No color styling)
    table_style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align text to center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for header
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Body font
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Body font size
    ])
    table.setStyle(table_style)

    # Add the table to elements
    elements.append(table)
    elements.append(Spacer(1, 24))

    # Footer: Requested by, Accountant, GM(A) with enough space for signatures
    footer_table_data = [
        ['Requested by: ', '', 'Accountant: ', '', 'GM(A): ']
    ]
    footer_table = Table(footer_table_data, colWidths=[80, 80, 80, 80, 80])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT')  # Align text to the left
    ]))
    elements.append(footer_table)

    # Build the PDF
    doc.build(elements)

    return response

#backup for Used print
def needbackup_print_pdf(request):
    # Create the HttpResponse object with the appropriate PDF headers
    response = HttpResponse(content_type='application/pdf')
    unique_reference_count = Need.objects.exclude(reference_id__isnull=True).values('reference_id').distinct().count()
    
    # Fetch data from your model
    entries = Need.objects.filter(is_acknowledged=True, is_printed=True)

    if not entries.exists():  
         messages.error(request,'There is No Data to Print!')# Check if there are no entries
        # Render an error message on a new HTML page
         return redirect(' backup_need_view')  # Return error template instead of PDF

    response['Content-Disposition'] = f'attachment; filename="rit/mess/need/{unique_reference_count + 1}.pdf"'

    # Create the PDF object
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    normal_style = styles['Normal']
    title_style = styles['Heading1']

    # Add the logo and institution name in a table to align them horizontally
    logo_path = r'C:\Users\sdhan\OneDrive\Desktop\Projects\Hostel mess\base\static\images\RIT_logo.png'  # Your logo's path
    logo = Image(logo_path)
    logo.drawHeight = 1.25 * inch  # Adjust height
    logo.drawWidth = 1.25 * inch  # Adjust width

    # Create a table with the logo and institution name side by side
    title_table_data = [
        [logo, Paragraph('RAMCO INSTITUTE OF TECHNOLOGY RAJAPALAYAM', title_style)]
    ]
    title_table = Table(title_table_data, colWidths=[1.5 * inch, 4.5 * inch])  # Adjust column widths as needed
    title_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),  # Vertically align logo in the middle
        ('VALIGN', (1, 0), (1, 0), 'MIDDLE'),  # Vertically align text in the middle
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 20))

    # Add placeholders for supplier and date
    elements.append(Paragraph('To : ___________                 Date : _______________ ', normal_style))
    elements.append(Spacer(1, 15))
    elements.append(Paragraph('       M/s  :  _________________________', normal_style))
    elements.append(Spacer(1, 15))

    # Instructional text
    elements.append(Paragraph('Please supply the following list of items:', normal_style))
    elements.append(Spacer(1, 12))

    # Define table data (headers)
    data = [['S.No.', 'Date', 'Name of the item', 'Required Quantity','unit']]

    count = 1
    reference_id = f"rit/mess/Need/{unique_reference_count + 1}"

    # Populate the table with data
    for entry in entries:
        data.append([
            count,
            str(entry.date or ''),  # Incremental number
            str(entry.name or ''),  # Name of the item
            str(entry.quantity if entry.quantity is not None else '')  ,# Quantity
            str(entry.unit or '')
        ])

        # Assign reference_id and mark as printed
        entry.reference_id = reference_id
        entry.is_printed = True
        entry.save()

        count += 1

    # Create the table without colors
    table = Table(data, colWidths=[80,120, 150,120,80])  # Adjust column widths to match

    # Add table style (No color styling)
    table_style = TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Align text to center
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold header
        ('FONTSIZE', (0, 0), (-1, 0), 12),  # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),  # Padding for header
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add grid lines
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),  # Body font
        ('FONTSIZE', (0, 1), (-1, -1), 10),  # Body font size
    ])
    table.setStyle(table_style)

    # Add the table to elements
    elements.append(table)
    elements.append(Spacer(1, 24))

    # Footer: Requested by, Accountant, GM(A) with enough space for signatures
    footer_table_data = [
        ['Requested by: ', '', 'Accountant: ', '', 'GM(A): ']
    ]
    footer_table = Table(footer_table_data, colWidths=[80, 80, 80, 80, 80])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT')  # Align text to the left
    ]))
    elements.append(footer_table)

    # Build the PDF
    doc.build(elements)

    return response




#Shopkeeper needed to be purchase(purchase.html) and approval to gm
@login_required
def materials_purchase_view(request):
    user_role = request.session.get('role', 'Guest')
    if request.method == 'POST':
        # Get all the submitted data as lists (one entry for each material)
        material_dates = request.POST.getlist('material-date')
        material_names = request.POST.getlist('material-name')
        material_quantity = request.POST.getlist('material-quantity')
        material_units = request.POST.getlist('material-unit')

        # Loop over the length of the materials and save each entry to the database
        for date, name, quantity, unit in zip(material_dates, material_names, material_quantity, material_units):
            if name and quantity:  # Ensure valid data before saving
                purchase = Purchase.objects.create(
                    name=name,  # Use the individual 'name' value
                    quantity=float(quantity),  # Convert the individual 'quantity' value to float
                    unit=unit,  # Use the individual 'unit' value
                    date=date if date else timezone.now().date()  # Use today's date if none is provided
                )
                purchase.save()

        email_subject = "Material Purchase Acknowledgement Required"
        email_message =  ("Sir,\n\n"
                         "Materials have been added for purchase and require your acknowledgement.\n\n"
                         "Please visit the page at your earliest convenience to review and acknowledge the purchase.\n\n"
                         "Thank you for your attention to this matter")
        
        # Define the recipient list
        recipient_list = ['sdhanesh09829@gmail.com']  # Change to the desired recipient email
        
        # Send the email
        send_mail(
            subject=email_subject,
            message=email_message,
            from_email='ehostelrit@gmail.com',  # Use the email you configured in settings
            recipient_list=recipient_list,
            fail_silently=False,
        )
        
        messages.success(request, 'Data saved and email sent successfully!')
        return redirect('materials_purchase_view')
    context = {
        'user_role': user_role,  # Add the user's role to the context
    }  # Redirect to the desired page

    return render(request, 'base/purchase.html',context)



#mess-incharge needed to use stock (need.html) and approve to gm
@login_required
def materials_need_view(request):
    user_role = request.session.get('role', 'Guest')
    
    if request.method == 'POST':
        # Get all the submitted data as lists (one entry for each material)
        material_dates = request.POST.getlist('material-date')
        No_of_persons=request.POST.getlist('No_of_persons')
        material_names = request.POST.getlist('material-name')
        material_quantity = request.POST.getlist('material-quantity')
        material_units = request.POST.getlist('material-unit')

        # Loop over the length of the materials and save each entry to the database
        for date,No_of_persons, name, quantity, unit in zip(material_dates,No_of_persons, material_names, material_quantity, material_units):
            if name and quantity:  # Ensure valid data before saving
                need = Need.objects.create(
                    No_of_persons=No_of_persons,
                    name=name,  # Use the individual 'name' value
                    quantity=float(quantity),  # Convert the individual 'quantity' value to float
                    unit=unit,  # Use the individual 'unit' value
                    date=date if date else timezone.now().date()  # Use today's date if none is provided
                )
                need.save()

        # Send an email after saving the data
        email_subject = "Material Need Acknowledgement Required"
        email_message = ("Sir,\n\n"
                         "Materials have been added as needed and require your acknowledgement.\n\n"
                         "Please visit the page at your earliest convenience to review and acknowledge the need.\n\n"
                         "Thank you for your attention to this matter")
        
        # Define the recipient list
        recipient_list = ['sdhanesh09829@gmail.com']  # Change to the desired recipient email

        # Send the email
        send_mail(
            subject=email_subject,
            message=email_message,
            from_email='ehostelrit@gmail.com',  # Use the email you configured in settings
            recipient_list=recipient_list,
            fail_silently=False,
        )

        # Show a success message and redirect to the need view page
        messages.success(request, 'Data saved and email sent successfully!')
        return redirect('materials_need_view')

    # Context for rendering the template
    context = {
        'user_role': user_role,
    }
    
    # Render the need.html page with the context
    return render(request, 'base/need.html', context)


#shopkeeper to store the received purchase material(receive.html)
def materials_receive_view(request):
    user_role = request.session.get('role', 'Guest')

    if request.method == "POST":
        # Get lists from the POST data
        names = request.POST.getlist('name')
        received_dates = request.POST.getlist('received_date')
        received_quantities = request.POST.getlist('received_quantity')
        units = request.POST.getlist('unit')
        remarks_list = request.POST.getlist('remarks')

        # Ensure all lists have the same length
        if not (len(names) == len(received_dates) == len(received_quantities) == len(units) == len(remarks_list)):
            return HttpResponse('Error: Form data mismatch')
        
        processed_purchases = set()  # Track processed purchases
        for name, received_date, received_quantity, unit, remarks in zip(names, received_dates, received_quantities, units, remarks_list):
    # Fetch the latest or most relevant Purchase record by name
            try:
                purchase = Purchase.objects.filter(name=name, is_acknowledged=True, is_printed=True).latest('date')
            except Purchase.DoesNotExist:
                return HttpResponse(f'Error: No acknowledged or printed purchase found for material "{name}".')

            # Parse received quantity
            received_quantity = Decimal(received_quantity) if received_quantity else None

            # Handle "not received" or pending entries for the same name and received date
            if received_quantity is not None:
                try:
                    existing_entry = Received.objects.get(name=name, received_date=received_date, received_quantity=0)
                    existing_entry.delete()
                except Received.DoesNotExist:
                    pass  # No "not received" entry found; continue

            # Create or update the Received entry
            received_entry, created = Received.objects.get_or_create(
                name=name,
                received_date=received_date,  # Use user-provided received_date
                defaults={
                    'quantity': purchase.quantity,  # Fetch from the latest purchase
                    'unit': unit if unit else purchase.unit,
                    'received_quantity': received_quantity,
                    'remarks': remarks if remarks else '',
                    'status': 'Pending' if received_quantity is None else (
                        'Received' if received_quantity >= purchase.quantity else 'Not Received'
                    ),
                }
            )

            # If the entry already exists, update it
            if not created:
                if received_quantity is not None:
                    received_entry.received_quantity = received_quantity
                    received_entry.status = 'Received' if received_quantity >= purchase.quantity else 'Pending'
                if remarks:
                    received_entry.remarks = remarks
                if unit:
                    received_entry.unit = unit
                received_entry.save()

          # Mark purchase as fetched if received data is successfully stored
            if purchase.id not in processed_purchases:
                purchase.is_fetched = True
                purchase.save()
                processed_purchases.add(purchase.id)  # Ensure we don't update multiple times for the same purchase

                

        return redirect('received_view')

    context = {
        'user_role': user_role,
        'purchases': Purchase.objects.all(),  # Pass available purchases to the template
    }
    return render(request, 'base/received.html', context)



# Fetch data based on selected purchase date receive.html
def fetch_purchase_data(request):
    purchase_date = request.GET.get('purchase_date')

    if not purchase_date:
        return JsonResponse({'error': 'Purchase date is required.'})
    
    # Check if any data exists in the Received model for the selected date
    received_data = Received.objects.filter(received_date=purchase_date)

    if received_data.exists():
          # Update statuses dynamically before proceeding
        for entry in received_data:
            if entry.received_quantity is None or entry.received_quantity == 0:
                entry.status = "Not Received"
            elif entry.received_quantity > 0 and entry.received_quantity < entry.quantity:
                entry.status = "Pending"
            elif entry.received_quantity >= entry.quantity:
                entry.status = "Received"
            entry.save()

        # Re-fetch the updated data to ensure consistency

        pending_data = received_data.filter(status="Pending")
        not_received_data = received_data.filter(status="Not Received")
        received_data_count = received_data.filter(status="Received").count()
          # All data marked as received
        if received_data_count == received_data.count():
            return JsonResponse({'message': 'All data for the selected date has already been stored!'})

     # Both pending and not received exist
        if pending_data.exists() and not_received_data.exists():
            return JsonResponse({
                'prompt': 'Both Pending and Not Received entries exist. Select between Upgrade or New Stock.',
                'pending': [{'id': entry.id, 'name': entry.name, 'unit': entry.unit, 'quantity': entry.quantity} for entry in pending_data],
                'not_received': [{'id': entry.id, 'name': entry.name, 'unit': entry.unit, 'quantity': entry.quantity} for entry in not_received_data],
            })
       
         # Only pending exists
        if pending_data.exists():
            return JsonResponse({'status': 'pending', 'data': [{'id': entry.id, 'name': entry.name, 'unit': entry.unit, 'quantity': entry.quantity} for entry in pending_data]})

        # Only not received exists
        if not_received_data.exists():
            return JsonResponse({'status': 'not_received', 'data': [{'id': entry.id, 'name': entry.name, 'unit': entry.unit, 'quantity': entry.quantity} for entry in not_received_data]})

 # Fetch data from the Purchase model if no entries exist in Received
    purchases = Purchase.objects.filter(date=purchase_date, is_acknowledged=True, is_printed=True, is_fetched=False)

    if purchases.exists():
        
        purchase_list = [{'id': purchase.id, 'name': purchase.name, 'quantity': purchase.quantity, 'unit': purchase.unit} for purchase in purchases]
        return JsonResponse({'status': 'new', 'purchases': purchase_list})
    

    return JsonResponse({'error': 'No unfetched purchases found for this date.'})


#fetch the data for data.html
def fetch_received_data(request):
    """
    Fetches only unfetched received data from the Received model based on the entered date.
    Ensures pending items are not fetched and already fetched data is not refetched.
    """
    date = request.GET.get('date')  # Get the date from the request

    if not date:
        return JsonResponse({'error': 'Invalid or missing date input.'}, status=400)

    try:
        # Check if there are pending items for the given date
        if Received.objects.filter(received_date=date, status="Pending").exists():
            return JsonResponse({'error': f"Some materials for {date} are still pending in the Received Entries."}, status=400)

        # Check if all received data for this date is already fetched
        if not Received.objects.filter(received_date=date, status="Received", is_fetched=False).exists():
            return JsonResponse({'error': f"Data for {date} already fetched ."}, status=400)

        # Fetch only received items that are NOT fetched
        received_items = Received.objects.filter(received_date=date, status="Received", is_fetched=False)

        if received_items.exists():
            materials = []
            for item in received_items:
                materials.append({
                    'id': item.id,  # ID for reference
                    'date': item.received_date.strftime('%Y-%m-%d'),
                    'name': item.name,
                    'quantity': item.received_quantity,
                    'unit': item.unit,
                })

            return JsonResponse({'materials': materials})

        return JsonResponse({'error': 'No new materials to fetch for the selected date.'}, status=404)

    except Exception as e:
        return JsonResponse({'error': f"An error occurred: {str(e)}"}, status=500)


#fetch the data using for the used.html
def fetch_materials_by_date(request):
    """
    Fetches materials from the Need model for a given date.
    Prevents fetching if materials have already been used.
    """
    if request.method == 'GET':
        selected_date = request.GET.get('date')

        if not selected_date:
            return JsonResponse({'error': 'Invalid or missing date input.'}, status=400)

        try:
            # Check if any data for this date has already been used
            if Need.objects.filter(date=selected_date, is_fetched=True).exists():
                return JsonResponse({'error': f"Data for {selected_date} has already been fetched and used."}, status=400)

            # Fetch only data that hasn't been used yet
            materials = Need.objects.filter(
                date=selected_date, 
                is_acknowledged=True, 
                is_printed=True, 
                is_fetched=False  # Ensure only unfetched data is retrieved
            )

            if materials.exists():
                material_list = [
                    {
                        'id': material.id,
                        'date': material.date.strftime('%Y-%m-%d'),
                        'name': material.name,
                        'quantity': material.quantity,
                        'unit': material.unit,
                    }
                    for material in materials
                ]
                return JsonResponse({'materials': material_list})

            return JsonResponse({'error': 'No new materials to fetch for this date.'}, status=404)

        except Exception as e:
            return JsonResponse({'error': f"An error occurred: {str(e)}"}, status=500)

        
#fetch the mess schedule for usage    
def get_mess_menu(request):
    if request.method == "GET":
        day = request.GET.get('day')
        session = request.GET.get('session')

        try:
            menu = Mess_menu_Schedule.objects.get(day=day)
            menu_data = getattr(menu, session, '')  # Dynamically get session data (e.g., breakfast, lunch)
            return JsonResponse({'menu': menu_data})
        except Mess_menu_Schedule.DoesNotExist:
            return JsonResponse({'menu': None})

    return JsonResponse({'error': 'Invalid request'}, status=400)


#page for viewing the received material(receive_view.html)
@login_required
def received_view(request):
    user_role = request.session.get('role', 'Guest')
    selected_names = request.GET.getlist('names')  # List of selected material names
    selected_month = request.GET.get('month', '')  # Selected month
    selected_date = request.GET.get('selected_date', '')  # Specific date

    # Get all received materials
    received_data = Received.objects.all()

    # Apply filters
    if selected_names:
        received_data = received_data.filter(name__in=selected_names)

    if selected_month:
        received_data = received_data.filter(received_date__month=selected_month)

    if selected_date:
        received_data = received_data.filter(received_date=selected_date)

    # Get distinct material names for the dropdown
    all_materials = Received.objects.values_list('name', flat=True).distinct()

    # Context for the template
    context = {
        'received_data': received_data,
        'all_materials': all_materials,
        'selected_names': selected_names,
        'selected_month': selected_month,
        'user_role': user_role,
        'selected_date': selected_date,
    }

    return render(request, 'base/received_view.html', context)

def logout_view(request):
    logout(request)  # This will log out the user
    return redirect('home')  # Redirect to home page after logout


#page for entering the new mess schedule(mess_entry.html)
@login_required
def mess_entry_view(request):
    user_role = request.session.get('role', 'Guest')
    if request.method == 'POST':
        day = request.POST.get('form-day')
        year = request.POST.get('form-year')
        breakfast = request.POST.get('form-breakfast')
        lunch = request.POST.get('form-lunch')
        snacks = request.POST.get('form-snacks')
        dinner = request.POST.get('form-dinner')

        print("Day:", day)
        print("Year:", year)
        print("Breakfast:", breakfast)
        print("Lunch:", lunch)
        print("Snacks:", snacks)
        print("Dinner:", dinner)

        if day and year and breakfast and lunch and snacks and dinner:
            menu = Mess_menu_Schedule(
                day=day,
                year=year,
                breakfast=breakfast,
                lunch=lunch,
                snacks=snacks,
                dinner=dinner
            )
            menu.save()
            messages.success(request, "Menu saved successfully!")
        else:
            messages.error(request, "All fields are required.")
        
    context = {
        'user_role': user_role,  # Add the user's role to the context
    }



    return render(request, 'base/mess entry.html',context)


#page for adding new material item(item_table.html)
@login_required
def add_item(request):
    user_role = request.session.get('role', 'Guest')
    
    if request.method == 'POST':
        # Handle item addition
        item_name = request.POST.get('item-name').strip()
        if item_name:
            if Item.objects.filter(name=item_name).exists():
                messages.error(request, "Item already exists.")
            else:
                Item.objects.create(name=item_name)
                messages.success(request, "Item added successfully.")

        # Handle file upload and process the file content
        uploaded_file = request.FILES.get('file_Upload')
        if uploaded_file:
            try:
                if uploaded_file.name.endswith('.csv'):
                    data_set = uploaded_file.read().decode('UTF-8')
                    io_string = io.StringIO(data_set)
                    for row in csv.reader(io_string, delimiter=','):
                        item_name = row[0].strip()
                        if not Item.objects.filter(name=item_name).exists():
                            Item.objects.create(name=item_name)
                    messages.success(request, "CSV file uploaded and items added successfully.")
                
                elif uploaded_file.name.endswith('.xlsx'):
                    workbook = load_workbook(uploaded_file)
                    sheet = workbook.active
                    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
                        item_name = row[0].strip() if row[0] else None
                        if item_name and not Item.objects.filter(name=item_name).exists():
                            Item.objects.create(name=item_name)
                    messages.success(request, "Excel file uploaded and items added successfully.")
                
                else:
                    messages.error(request, "Invalid file format. Please upload a CSV or Excel file.")
            
            except Exception as e:
                messages.error(request, f"Error processing the file: {str(e)}")

        # Redirect to display the updated item list
        return redirect('add_item')

    # Fetch all items to display in the table
    items = Item.objects.all()
    return render(request, 'base/add_item.html', {'items': items, 'user_role': user_role})

#suggestion for all materials
def get_item_suggestions(request):
    if 'term' in request.GET:
        qs = Item.objects.filter(name__icontains=request.GET['term'])  # Filter based on input
        names = [item.name for item in qs]
        return JsonResponse(names, safe=False)
    return JsonResponse([], safe=False)

#backup purchase view(backup_purchase.html)
@login_required
def backup_purchase_view(request):
    user_role = request.session.get('role', 'Guest')
    data_entries = Purchase.objects.all()

    # Handle filters
    selected_names = request.GET.getlist('names')
    selected_month = request.GET.get('month')
    selected_date = request.GET.get('selected_date')

    if selected_names:
        data_entries = data_entries.filter(name__in=selected_names)
    if selected_month:
        data_entries = data_entries.filter(date__month=selected_month)
    if selected_date:
        data_entries = data_entries.filter(date=selected_date)

    # Pass the filtered data to the template
    context = {
        'data_entries': data_entries,
        'all_materials': Purchase.objects.values_list('name', flat=True).distinct(),
        'selected_names': selected_names,
        'selected_month': selected_month,
        'selected_date': selected_date,
        'user_role':user_role,
    }

    return render(request, 'base/backup_purchase.html', context)


#backup need view(backup_need.html)
@login_required
def backup_need_view(request):
    user_role = request.session.get('role', 'Guest')
    # Get filter criteria from request
    selected_names = request.GET.getlist('names')
    selected_month = request.GET.get('month')
    selected_date = request.GET.get('selected_date')

    # Filter data based on selected criteria
    needs = Need.objects.all()

    if selected_names:
        needs = needs.filter(name__in=selected_names)
    if selected_month:
        needs = needs.filter(date__month=selected_month)
    if selected_date:
        needs = needs.filter(date=selected_date)

    context = {
        'needs': needs,
        'all_names': Need.objects.values_list('name', flat=True).distinct(),
        'selected_names': selected_names,
        'selected_month': selected_month,
        'selected_date': selected_date,
        'user_role':user_role,
    }
    return render(request, 'base/backup_need.html', context)


# to visualize the stock (data_visualization.html)
@login_required
def visualization(request):
    user_role = request.session.get('role', 'Guest')
    materials = Material.objects.values('name').distinct()
    months = [{'value': i, 'name': datetime(2023, i, 1).strftime('%B')} for i in range(1, 13)]
    
    context = {
        'materials': materials,
        'months': months,
        'user_role':user_role,
    }
    return render(request, 'base/data_visualization.html', context)


#filter the visualization 
def update_charts(request):
    material_names = request.GET.getlist('material_names')
    quantity_type = request.GET.get('quantity_type', 'available_quantity')
    month = request.GET.get('month', '')
    date = request.GET.get('date', '')

    # Filter the queryset as needed
    queryset = Material.objects.all()
    if material_names:
        queryset = queryset.filter(name__in=material_names)
    if month:
        queryset = queryset.filter(date__month=month)

    if date:
        # When date is selected, filter entries for that specific date
        queryset = queryset.filter(date=date)
    else:
        # Find the latest entry for each material when no date is specified
        queryset = queryset.filter(
            date=Subquery(
                Material.objects.filter(name=OuterRef('name'))
                .order_by('-date')
                .values('date')[:1]
            )
        )

    # Aggregate data for selected quantity type
    data = queryset.values('name').annotate(total_quantity=Sum(quantity_type))

    # Prepare professional color palette
    palette = sns.color_palette("Set2", len(data))
    colors = [to_hex(color) for color in palette]

    # Prepare data for bar chart
    bar_chart_data = {
        'x': [item['name'] for item in data],
        'y': [item['total_quantity'] for item in data],
        'type': 'bar',
        'marker': {'color': colors}
    }

    bar_chart_layout = {
        'xaxis': {'title': 'Material Name'},
        'yaxis': {'title': f'{quantity_type.capitalize()} in kg/L'},
        'title': 'Material Quantity Visualization'
    }

    # Prepare data for pie chart
    pie_chart_data = {
        'labels': [item['name'] for item in data],
        'values': [item['total_quantity'] for item in data],
        'type': 'pie',
        'marker': {'colors': colors}
    }

    return JsonResponse({
        'bar_chart': [bar_chart_data],
        'bar_chart_layout': bar_chart_layout,
        'pie_chart': [pie_chart_data]
    })

